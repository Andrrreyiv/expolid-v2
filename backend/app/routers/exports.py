import io
from datetime import datetime

from fastapi import APIRouter, Depends, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.deps import get_current_user
from app.models import Contact, Exhibition, User

router = APIRouter(prefix="/api/exports", tags=["exports"])

# (заголовок, источник). Источник — имя атрибута либо callable(contact)->value.
COLUMNS: list[tuple[str, object]] = [
    ("ID", "id"),
    ("Создан", "created_at"),
    ("Связь", lambda c: "визитка владельца" if not c.card_belongs_to_other else "визитка коллеги/другого"),
    ("Имя", "name"),
    ("Компания", "company"),
    ("Должность", "position"),
    ("Email", "email"),
    ("Телефон", "phone"),
    ("Сайт", "website"),
    ("Telegram", "telegram"),
    ("WhatsApp", "whatsapp"),
    ("LinkedIn", "linkedin"),
    ("Павильон", "pavilion"),
    ("Стенд", "stand"),
    ("Тип", "contact_type"),
    ("Статус", "status"),
    ("Менеджер", lambda c: c.assignee.name if c.assignee else ""),
    ("Кто записал", lambda c: c.captured_by.name if c.captured_by else ""),
    ("AI-скор 1–100", "ai_score"),
    ("Причина скоринга", "ai_score_reason"),
    ("Заметка", "note"),
    ("Имя на визитке", "card_owner_name"),
    ("Должность на визитке", "card_owner_position"),
    ("Email на визитке", "card_owner_email"),
    ("Телефон на визитке", "card_owner_phone"),
]


@router.get("/contacts.xlsx")
async def export_contacts(
    exhibition_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    stmt = (
        select(Contact)
        .where(Contact.organization_id == user.organization_id)
        .options(selectinload(Contact.assignee), selectinload(Contact.captured_by))
    )
    if exhibition_id is not None:
        stmt = stmt.where(Contact.exhibition_id == exhibition_id)
    stmt = stmt.order_by(Contact.created_at.desc())
    res = await db.execute(stmt)
    contacts = list(res.scalars().all())

    exhibition_name: str | None = None
    if exhibition_id is not None:
        exh = await db.get(Exhibition, exhibition_id)
        if exh is not None and exh.organization_id == user.organization_id:
            exhibition_name = exh.name

    wb = Workbook()
    ws = wb.active
    ws.title = "Контакты"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    for idx, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, c in enumerate(contacts, start=2):
        for col_idx, (_, source) in enumerate(COLUMNS, start=1):
            if callable(source):
                value = source(c)
            else:
                value = getattr(c, source, None)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M")
            elif isinstance(value, bool):
                value = "да" if value else ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    # crude column width
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    suffix = exhibition_name.replace(" ", "_") if exhibition_name else "all"
    filename = f"expolid_{suffix}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
